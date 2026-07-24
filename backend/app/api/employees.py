"""
Employee API endpoints.
"""

import csv
import io
import logging
import re
from urllib.parse import quote
from typing import Optional
from datetime import date, datetime
from fastapi import APIRouter, Depends, HTTPException, Query, Header, UploadFile, File
from fastapi.responses import Response
from sqlalchemy.orm import Session
from sqlalchemy import and_, exists, func, literal, or_
from app.core.database import get_db
from app.models.employee import Employee, EmployeeAuditLog, EmployeePerformanceSnapshot
from app.models.leave_attendance import LeaveBalance, LeaveRequest
from app.models.operations import ActionInboxItem, Allocation, Notification, Project
from app.models.training import TrainingEnrollment
from app.schemas.employee import AddEmployeeRequest, AddEmployeeResponse, UpdateEmployeeRequest
from app.services.employee_service import create_employee
from app.services.audit_service import changed_fields, log_audit, log_authorization_failure
from app.services.settings_service import get_current_employee, is_admin_role, normalize_role, require_admin_employee
from app.services.security_service import (
    export_employee_csv,
    log_sensitive_access,
    require_export_level,
)
import base64

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/employees", tags=["Employees"])

BULK_EMPLOYEE_HEADERS = (
    "first_name", "last_name", "work_email", "phone", "country_code",
    "department", "designation", "role", "reporting_manager",
    "workforce_type", "work_arrangement", "work_city", "work_state",
    "work_country", "joining_date",
)
BULK_REQUIRED_FIELDS = (
    "first_name", "last_name", "work_email", "phone", "department", "role",
    "reporting_manager", "workforce_type", "work_arrangement", "joining_date",
)
BULK_ALLOWED_ROLES = {"super_admin", "hr_admin", "manager", "employee", "trainee"}
BULK_ALLOWED_ARRANGEMENTS = {"remote", "hybrid", "office", "onshore", "offshore"}


def _normalize_bulk_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower()).strip("_")


def _bulk_cell_text(value: object) -> str:
    if isinstance(value, (date, datetime)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    return str(value or "").strip()


def _parse_bulk_employee_file(filename: str, content: bytes) -> list[dict[str, str]]:
    suffix = (filename or "").lower().rsplit(".", 1)[-1]
    if suffix == "csv":
        try:
            text_content = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise HTTPException(status_code=400, detail="CSV files must use UTF-8 encoding.") from exc
        reader = csv.DictReader(io.StringIO(text_content))
        if not reader.fieldnames:
            raise HTTPException(status_code=400, detail="The uploaded CSV has no header row.")
        rows = [
            {_normalize_bulk_header(key): str(value or "").strip() for key, value in row.items() if key}
            for row in reader
            if any(str(value or "").strip() for value in row.values())
        ]
    elif suffix == "xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise HTTPException(status_code=500, detail="Excel import support is not installed.") from exc
        try:
            workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            values = sheet.iter_rows(values_only=True)
            headers = [_normalize_bulk_header(value) for value in next(values, ())]
            if not any(headers):
                raise HTTPException(status_code=400, detail="The uploaded workbook has no header row.")
            rows = []
            for values_row in values:
                row = {headers[index]: _bulk_cell_text(value) for index, value in enumerate(values_row) if index < len(headers) and headers[index]}
                if any(row.values()):
                    rows.append(row)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=400, detail="The Excel workbook could not be read.") from exc
    else:
        raise HTTPException(status_code=400, detail="Upload a .csv or .xlsx file.")
    if not rows:
        raise HTTPException(status_code=400, detail="The uploaded file contains no employee rows.")
    if len(rows) > 500:
        raise HTTPException(status_code=400, detail="Bulk import supports up to 500 employee rows at a time.")
    return rows


def _validate_bulk_employee_rows(db: Session, rows: list[dict[str, str]]) -> list[dict]:
    employees = db.query(Employee).all()
    existing_emails = {(employee.work_email or "").strip().lower() for employee in employees}
    departments = {(employee.department or "").strip().lower() for employee in employees if (employee.department or "").strip()}
    managers = {
        (employee.work_email or "").strip().lower()
        for employee in employees
        if (employee.work_email or "").strip()
    } | {
        employee_name(employee).strip().lower()
        for employee in employees
        if employee_name(employee).strip()
    }
    seen_emails: set[str] = set()
    validated = []
    for index, raw in enumerate(rows, start=2):
        row = {header: (raw.get(header) or "").strip() for header in BULK_EMPLOYEE_HEADERS}
        row["work_email"] = row["work_email"].lower()
        row["role"] = row["role"].lower().replace(" ", "_").replace("-", "_")
        errors: list[str] = []
        missing = [field for field in BULK_REQUIRED_FIELDS if not row[field]]
        if missing:
            errors.append(f"{missing[0].replace('_', ' ').title()} required")
        if row["work_email"] and not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", row["work_email"]):
            errors.append("Invalid email")
        elif row["work_email"] in existing_emails or row["work_email"] in seen_emails:
            errors.append("Duplicate email")
        if row["work_email"]:
            seen_emails.add(row["work_email"])
        if row["department"] and departments and row["department"].lower() not in departments:
            errors.append("Unknown department")
        if row["role"] and row["role"] not in BULK_ALLOWED_ROLES:
            errors.append("Unknown role")
        if row["reporting_manager"] and row["reporting_manager"].lower() not in managers:
            errors.append("Manager not found")
        arrangement = row["work_arrangement"].lower()
        if arrangement and arrangement not in BULK_ALLOWED_ARRANGEMENTS:
            errors.append("Invalid work arrangement")
        parsed_date = None
        if row["joining_date"]:
            for date_format in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
                try:
                    parsed_date = datetime.strptime(row["joining_date"], date_format).date()
                    break
                except ValueError:
                    pass
            if parsed_date is None:
                errors.append("Invalid joining date")
        payload = {
            "first_name": row["first_name"], "last_name": row["last_name"],
            "work_email": row["work_email"], "phone": row["phone"],
            "country_code": row["country_code"] or "+1",
            "department": row["department"], "designation": row["designation"] or None,
            "role": row["role"], "reporting_manager": row["reporting_manager"],
            "workforce_type": row["workforce_type"],
            "work_location": row["work_arrangement"],
            "work_city": row["work_city"] or None, "work_state": row["work_state"] or None,
            "work_country": row["work_country"] or None,
            "joining_date": parsed_date.isoformat() if parsed_date else row["joining_date"],
        }
        validated.append({
            "row": index,
            "name": f"{row['first_name']} {row['last_name']}".strip() or "Unnamed employee",
            "email": row["work_email"],
            "department": row["department"],
            "valid": not errors,
            "error": errors[0] if errors else None,
            "errors": errors,
            "payload": payload,
        })
    return validated


SELF_PROFILE_FIELDS = {
    "personal_email",
    "phone",
    "country_code",
    "date_of_birth",
    "gender",
    "emergency_contact_name",
    "emergency_contact_phone",
    "emergency_contact_relation",
    "current_address",
    "permanent_address",
    "profile_image_url",
}

ADMIN_EMPLOYEE_FIELDS = SELF_PROFILE_FIELDS | {
    "first_name",
    "last_name",
    "department",
    "designation",
    "role",
    "workforce_type",
    "workforce_status",
    "employment_status",
    "work_location",
    "work_city",
    "work_state",
    "work_country",
    "reporting_manager",
    "joining_date",
    "date_of_exit",
    "inactive_reason",
    "onboarding_type",
    "notes",
}

SENSITIVE_EMPLOYMENT_FIELDS = {
    "department",
    "designation",
    "role",
    "workforce_type",
    "workforce_status",
    "employment_status",
    "work_location",
    "work_city",
    "work_state",
    "work_country",
    "reporting_manager",
    "joining_date",
    "date_of_exit",
    "inactive_reason",
    "onboarding_type",
}


def serialize_employee(emp: Employee) -> dict:
    """Serialize employee data used by profile and employee screens."""
    return {
        "id": emp.id,
        "first_name": emp.first_name,
        "last_name": emp.last_name,
        "work_email": emp.work_email,
        "personal_email": emp.personal_email,
        "phone": emp.phone,
        "country_code": emp.country_code,
        "date_of_birth": str(emp.date_of_birth) if emp.date_of_birth else None,
        "gender": emp.gender,
        "department": emp.department,
        "designation": emp.designation,
        "role": emp.role,
        "workforce_type": emp.workforce_type,
        "workforce_status": emp.workforce_status,
        "employment_status": emp.employment_status,
        "work_location": emp.work_location,
        "work_city": emp.work_city,
        "work_state": emp.work_state,
        "work_country": emp.work_country,
        "location": emp.location,
        "joining_date": str(emp.joining_date) if emp.joining_date else None,
        "reporting_manager": emp.reporting_manager,
        "onboarding_type": emp.onboarding_type,
        "profile_image_url": emp.profile_image_url,
        "emergency_contact_name": emp.emergency_contact_name,
        "emergency_contact_phone": emp.emergency_contact_phone,
        "emergency_contact_relation": emp.emergency_contact_relation,
        "current_address": emp.current_address,
        "permanent_address": emp.permanent_address,
        "is_active": emp.is_active,
        "is_first_login": emp.is_first_login,
        "last_login_at": str(emp.last_login_at) if emp.last_login_at else None,
        "last_active_at": str(emp.last_active_at) if emp.last_active_at else None,
        "access_level": emp.access_level,
        "mfa_enabled": emp.mfa_enabled,
        "device_assigned": emp.device_assigned,
        "setup_code": emp.setup_code,
        "created_at": str(emp.created_at),
        "last_updated_at": str(emp.last_updated_at) if emp.last_updated_at else None,
        "updated_by": emp.updated_by,
    }


def employee_name(emp: Employee) -> str:
    return " ".join(part.strip() for part in [emp.first_name, emp.last_name] if part and part.strip())


def normalize_person_reference(value: str | None) -> str:
    return " ".join(re.findall(r"[a-z0-9]+", (value or "").lower()))


def resolve_manager_id(employee: Employee, employees: list[Employee]) -> str | None:
    by_id = {row.id: row for row in employees}
    if employee.manager_id and employee.manager_id != employee.id and employee.manager_id in by_id:
        return employee.manager_id
    reference = normalize_person_reference(employee.reporting_manager)
    if not reference or reference in {"self", "none", "not assigned", "unassigned"}:
        return None
    exact_matches = [
        row for row in employees
        if row.id != employee.id and reference in {
            normalize_person_reference(employee_name(row)),
            normalize_person_reference(row.work_email),
        }
    ]
    if len(exact_matches) == 1:
        return exact_matches[0].id
    reference_tokens = set(reference.split())
    partial_matches = [
        row for row in employees
        if row.id != employee.id and reference_tokens and reference_tokens.issubset(
            set(normalize_person_reference(employee_name(row)).split())
        )
    ]
    return partial_matches[0].id if len(partial_matches) == 1 else None


def creates_reporting_cycle(db: Session, employee_id: str, manager_id: str) -> bool:
    employees = db.query(Employee).all()
    manager_ids = {employee.id: resolve_manager_id(employee, employees) for employee in employees}
    manager_ids[employee_id] = manager_id
    current_id: str | None = manager_id
    visited: set[str] = set()
    while current_id and current_id not in visited:
        if current_id == employee_id:
            return True
        visited.add(current_id)
        current_id = manager_ids.get(current_id)
    return False


def stabilize_manager_map(employees: list[Employee], manager_ids: dict[str, str | None]) -> dict[str, str | None]:
    """Break legacy reporting cycles while retaining the most connected leader in each cycle."""
    stable = dict(manager_ids)
    employee_ids = {employee.id for employee in employees}
    root = next((employee.id for employee in employees if normalize_role(employee.role) == "super_admin"), None)
    for employee_id, manager_id in list(stable.items()):
        if manager_id not in employee_ids or manager_id == employee_id:
            stable[employee_id] = None

    while True:
        cycle: list[str] | None = None
        for start_id in employee_ids:
            path: list[str] = []
            positions: dict[str, int] = {}
            current_id: str | None = start_id
            while current_id and current_id in employee_ids:
                if current_id in positions:
                    cycle = path[positions[current_id]:]
                    break
                positions[current_id] = len(path)
                path.append(current_id)
                current_id = stable.get(current_id)
            if cycle:
                break
        if not cycle:
            return stable
        inbound_counts = {
            employee_id: sum(1 for manager_id in stable.values() if manager_id == employee_id)
            for employee_id in cycle
        }
        leader_id = max(cycle, key=lambda employee_id: (inbound_counts[employee_id], employee_id))
        stable[leader_id] = root if root and root not in cycle and root != leader_id else None


def can_view_employee_detail(actor: Employee, target: Employee) -> bool:
    if actor.id == target.id or is_admin_role(actor.role):
        return True
    if normalize_role(actor.role) != "manager":
        return False
    actor_name = employee_name(actor)
    return bool(
        target.manager_id == actor.id
        or target.reporting_manager == actor_name
        or target.reporting_manager == actor.work_email
    )


def changed_by_value(email: str | None, name: str | None, user_id: str | None) -> str:
    return email or name or user_id or "unknown"


def stringify_audit_value(value) -> str | None:
    if value is None:
        return None
    return str(value)


def log_employee_changes(
    db: Session,
    employee_id: str,
    old_values: dict,
    updates: dict,
    changed_by: str,
):
    important_fields = {
        "role",
        "department",
        "designation",
        "reporting_manager",
        "work_location",
        "access_level",
        "employment_status",
        "workforce_type",
        "workforce_status",
        "mfa_enabled",
        "device_assigned",
        "joining_date",
        "date_of_exit",
        "inactive_reason",
        "onboarding_type",
        "work_city",
        "work_state",
        "work_country",
    }

    for field in important_fields.intersection(updates.keys()):
        old_value = stringify_audit_value(old_values.get(field))
        new_value = stringify_audit_value(updates.get(field))
        if old_value == new_value:
            continue

        db.add(EmployeeAuditLog(
            employee_id=employee_id,
            action_type="field_changed",
            field_name=field,
            old_value=old_value,
            new_value=new_value,
            changed_by=changed_by,
            changed_at=datetime.utcnow(),
        ))


@router.get("/")
async def list_employees(
    search: Optional[str] = Query(None, description="Search by name or email"),
    department: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    role: Optional[str] = Query(None),
    location: Optional[str] = Query(None),
    project_status: Optional[str] = Query(None, pattern="^(in_project|bench|trainee)$"),
    reporting_manager: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user_id: str | None = Header(None, alias="x-user-id"),
    current_user_email: str | None = Header(None, alias="x-user-email"),
):
    """List employees with search, filters, and pagination."""
    require_admin_employee(db, current_user_id, current_user_email)
    query = db.query(Employee).filter(Employee.work_email != "superadmin@reknew.ai")
    organization_employees = query.all()
    organization_total = len(organization_employees)
    organization_ids = [employee.id for employee in organization_employees]
    stats_today = date.today()
    organization_allocated_ids = {
        employee_id for (employee_id,) in db.query(Allocation.employee_id).filter(
            Allocation.employee_id.in_(organization_ids),
            Allocation.project_id.isnot(None),
            Allocation.status == "active",
            Allocation.start_date <= stats_today,
            or_(Allocation.end_date.is_(None), Allocation.end_date >= stats_today),
        ).distinct().all()
    } if organization_ids else set()

    def is_trainee(employee: Employee) -> bool:
        role_name = (employee.role or "").strip().lower().replace(" ", "_").replace("-", "_")
        workforce_name = (employee.workforce_type or "").strip().lower().replace(" ", "_").replace("-", "_")
        return role_name == "trainee" or "trainee" in workforce_name

    trainee_ids = {employee.id for employee in organization_employees if is_trainee(employee)}
    in_project_ids = organization_allocated_ids - trainee_ids
    bench_ids = set(organization_ids) - trainee_ids - in_project_ids
    organization_stats = {
        "total": organization_total,
        "active": sum(1 for employee in organization_employees if (employee.employment_status or "").lower() == "active" and employee.is_active),
        "bench": len(bench_ids),
        "in_project": len(in_project_ids),
        "trainees": len(trainee_ids),
    }
    manager_references = db.query(Employee.manager_id, Employee.reporting_manager).filter(
        Employee.work_email != "superadmin@reknew.ai"
    ).all()
    referenced_manager_ids = {manager_id for manager_id, _ in manager_references if manager_id}
    referenced_managers = {
        manager.id: employee_name(manager)
        for manager in db.query(Employee).filter(Employee.id.in_(referenced_manager_ids)).all()
    } if referenced_manager_ids else {}
    reporting_manager_options = sorted({
        referenced_managers.get(manager_id) or (legacy_name or "").strip()
        for manager_id, legacy_name in manager_references
        if referenced_managers.get(manager_id) or (legacy_name or "").strip()
    }, key=str.lower)

    # Search
    if search:
        normalized_search = " ".join(search.strip().split())
        if normalized_search:
            full_name = (
                func.coalesce(Employee.first_name, "")
                + literal(" ")
                + func.coalesce(Employee.last_name, "")
            )
            searchable_fields = [
                Employee.first_name,
                Employee.last_name,
                Employee.work_email,
                Employee.department,
                Employee.designation,
                Employee.role,
                Employee.work_location,
                Employee.work_city,
                Employee.work_state,
                Employee.work_country,
                Employee.location,
                full_name,
            ]
            token_filters = []
            for token in normalized_search.split(" "):
                token_term = f"%{token}%"
                token_filters.append(or_(*[field.ilike(token_term) for field in searchable_fields]))

            phrase_term = f"%{normalized_search}%"
            query = query.filter(or_(
                full_name.ilike(phrase_term),
                Employee.work_email.ilike(phrase_term),
                and_(*token_filters),
            ))

    # Filters
    if department:
        query = query.filter(Employee.department.ilike(department))
    if status:
        query = query.filter(Employee.employment_status.ilike(status))
    if role:
        normalized_role = role.strip().lower().replace(" ", "_").replace("-", "_")
        role_value = func.lower(func.replace(func.replace(Employee.role, " ", "_"), "-", "_"))
        query = query.filter(role_value == normalized_role)
    if location:
        query = query.filter(Employee.work_location.ilike(location))
    if reporting_manager:
        normalized_manager = reporting_manager.strip()
        matching_manager_ids = [
            manager_id for manager_id, manager_name in referenced_managers.items()
            if manager_name.lower() == normalized_manager.lower()
        ]
        manager_filters = [Employee.reporting_manager.ilike(normalized_manager)]
        if matching_manager_ids:
            manager_filters.append(Employee.manager_id.in_(matching_manager_ids))
        query = query.filter(or_(*manager_filters))
    if project_status:
        today = date.today()
        trainee_expression = or_(
            func.lower(func.replace(func.replace(Employee.role, " ", "_"), "-", "_")) == "trainee",
            func.lower(Employee.workforce_type).like("%trainee%"),
        )
        active_project_exists = exists().where(and_(
            Allocation.employee_id == Employee.id,
            Allocation.project_id.isnot(None),
            Allocation.status == "active",
            Allocation.start_date <= today,
            or_(Allocation.end_date.is_(None), Allocation.end_date >= today),
        ))
        if project_status == "trainee":
            query = query.filter(trainee_expression)
        elif project_status == "in_project":
            query = query.filter(~trainee_expression, active_project_exists)
        else:
            query = query.filter(~trainee_expression, ~active_project_exists)

    # Count total before pagination
    total = query.count()

    # Pagination
    employees = query.order_by(Employee.created_at.desc()).offset(
        (page - 1) * per_page
    ).limit(per_page).all()

    employee_ids = [employee.id for employee in employees]
    today = date.today()
    allocated_employee_ids = {
        employee_id
        for (employee_id,) in db.query(Allocation.employee_id).filter(
            Allocation.employee_id.in_(employee_ids),
            Allocation.project_id.isnot(None),
            Allocation.status == "active",
            Allocation.start_date <= today,
            or_(Allocation.end_date.is_(None), Allocation.end_date >= today),
        ).distinct().all()
    } if employee_ids else set()
    manager_ids = {employee.manager_id for employee in employees if employee.manager_id}
    manager_names = {
        manager.id: employee_name(manager)
        for manager in db.query(Employee).filter(Employee.id.in_(manager_ids)).all()
    } if manager_ids else {}

    def project_status(employee: Employee) -> str:
        normalized_role = (employee.role or "").strip().lower().replace(" ", "_").replace("-", "_")
        normalized_workforce_type = (employee.workforce_type or "").strip().lower().replace(" ", "_").replace("-", "_")
        if normalized_role == "trainee" or "trainee" in normalized_workforce_type:
            return "trainee"
        return "in_project" if employee.id in allocated_employee_ids else "bench"

    return {
        "employees": [
            {
                "id": emp.id,
                "first_name": emp.first_name,
                "last_name": emp.last_name,
                "work_email": emp.work_email,
                "phone": emp.phone,
                "country_code": emp.country_code,
                "department": emp.department,
                "designation": emp.designation,
                "role": emp.role,
                "workforce_type": emp.workforce_type,
                "employment_status": emp.employment_status,
                "work_location": emp.work_location,
                "work_city": emp.work_city,
                "work_state": emp.work_state,
                "work_country": emp.work_country,
                "joining_date": str(emp.joining_date) if emp.joining_date else None,
                "reporting_manager": manager_names.get(emp.manager_id) or emp.reporting_manager or "Not assigned",
                "project_status": project_status(emp),
                "profile_image_url": emp.profile_image_url,
                "is_active": emp.is_active,
                "is_first_login": emp.is_first_login,
                "setup_code": emp.setup_code,
                "created_at": str(emp.created_at),
            }
            for emp in employees
        ],
        "total": total,
        "organization_total": organization_total,
        "stats": organization_stats,
        "reporting_manager_options": reporting_manager_options,
        "page": page,
        "per_page": per_page,
        "total_pages": (total + per_page - 1) // per_page,
    }


@router.get("/export")
async def export_employees(
    search: Optional[str] = Query(None, description="Search by name or email"),
    department: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    role: Optional[str] = Query(None),
    location: Optional[str] = Query(None),
    project_status: Optional[str] = Query(None, pattern="^(in_project|bench|trainee)$"),
    reporting_manager: Optional[str] = Query(None),
    level: str = Query("basic", pattern="^(basic|hr|payroll)$"),
    db: Session = Depends(get_db),
    current_user_id: str | None = Header(None, alias="x-user-id"),
    current_user_email: str | None = Header(None, alias="x-user-email"),
):
    """Export employees with server-side role checks and sensitive-access audit."""
    actor = require_admin_employee(db, current_user_id, current_user_email)
    require_export_level(actor, level)

    query = db.query(Employee).filter(Employee.work_email != "superadmin@reknew.ai")
    if search:
        normalized_search = " ".join(search.strip().split())
        if normalized_search:
            full_name = (
                func.coalesce(Employee.first_name, "")
                + literal(" ")
                + func.coalesce(Employee.last_name, "")
            )
            searchable_fields = [
                Employee.first_name,
                Employee.last_name,
                Employee.work_email,
                Employee.department,
                Employee.designation,
                Employee.role,
                Employee.work_location,
                Employee.work_city,
                Employee.work_state,
                Employee.work_country,
                Employee.location,
                full_name,
            ]
            token_filters = []
            for token in normalized_search.split(" "):
                token_term = f"%{token}%"
                token_filters.append(or_(*[field.ilike(token_term) for field in searchable_fields]))
            phrase_term = f"%{normalized_search}%"
            query = query.filter(or_(
                full_name.ilike(phrase_term),
                Employee.work_email.ilike(phrase_term),
                and_(*token_filters),
            ))
    if department:
        query = query.filter(Employee.department.ilike(department))
    if status:
        query = query.filter(Employee.employment_status.ilike(status))
    if role:
        normalized_role = role.strip().lower().replace(" ", "_").replace("-", "_")
        role_value = func.lower(func.replace(func.replace(Employee.role, " ", "_"), "-", "_"))
        query = query.filter(role_value == normalized_role)
    if location:
        query = query.filter(Employee.work_location.ilike(location))
    if reporting_manager:
        normalized_manager = reporting_manager.strip()
        manager_name = (
            func.coalesce(Employee.first_name, "")
            + literal(" ")
            + func.coalesce(Employee.last_name, "")
        )
        matching_manager_ids = db.query(Employee.id).filter(manager_name.ilike(normalized_manager)).all()
        manager_filters = [Employee.reporting_manager.ilike(normalized_manager)]
        if matching_manager_ids:
            manager_filters.append(Employee.manager_id.in_([row[0] for row in matching_manager_ids]))
        query = query.filter(or_(*manager_filters))
    if project_status:
        today = date.today()
        trainee_expression = or_(
            func.lower(func.replace(func.replace(Employee.role, " ", "_"), "-", "_")) == "trainee",
            func.lower(Employee.workforce_type).like("%trainee%"),
        )
        active_project_exists = exists().where(and_(
            Allocation.employee_id == Employee.id,
            Allocation.project_id.isnot(None),
            Allocation.status == "active",
            Allocation.start_date <= today,
            or_(Allocation.end_date.is_(None), Allocation.end_date >= today),
        ))
        if project_status == "trainee":
            query = query.filter(trainee_expression)
        elif project_status == "in_project":
            query = query.filter(~trainee_expression, active_project_exists)
        else:
            query = query.filter(~trainee_expression, ~active_project_exists)

    employees = query.order_by(Employee.created_at.desc()).limit(10000).all()
    log_sensitive_access(
        db,
        actor,
        action="employee_csv_export",
        target_type="employees",
        target_id=None,
        sensitivity_level="restricted" if level == "basic" else "confidential",
        reason=f"{level} employee export",
        metadata={
            "level": level,
            "row_count": len(employees),
            "filters": {
                "search": search,
                "department": department,
                "status": status,
                "role": role,
                "location": location,
                "project_status": project_status,
                "reporting_manager": reporting_manager,
            },
        },
    )
    db.commit()

    csv_content = export_employee_csv(employees, level)
    filename = quote(f"reknew-employees-{level}-{date.today().isoformat()}.csv")
    return Response(
        content=csv_content,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/bulk-template.csv")
async def bulk_employee_template(
    db: Session = Depends(get_db),
    current_user_id: str | None = Header(None, alias="x-user-id"),
    current_user_email: str | None = Header(None, alias="x-user-email"),
):
    """Download the canonical employee-import template."""
    require_admin_employee(db, current_user_id, current_user_email)
    output = io.StringIO()
    writer = csv.writer(output, lineterminator="\n")
    writer.writerow(BULK_EMPLOYEE_HEADERS)
    writer.writerow([
        "Asha", "Rao", "asha.rao@example.com", "5550101234", "+1",
        "Engineering", "Software Engineer", "employee", "Manager Name",
        "Full-Time Employee", "Remote", "Hartford", "CT", "USA", "2026-08-01",
    ])
    return Response(
        content=output.getvalue(),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="orbit-employee-import-template.csv"'},
    )


@router.post("/bulk/validate")
async def validate_bulk_employees(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user_id: str | None = Header(None, alias="x-user-id"),
    current_user_email: str | None = Header(None, alias="x-user-email"),
):
    """Validate a bulk employee file without writing any employee records."""
    require_admin_employee(db, current_user_id, current_user_email)
    content = await file.read()
    rows = _parse_bulk_employee_file(file.filename or "", content)
    validated = _validate_bulk_employee_rows(db, rows)
    return {
        "filename": file.filename,
        "row_count": len(validated),
        "ready": sum(1 for row in validated if row["valid"]),
        "invalid": sum(1 for row in validated if not row["valid"]),
        "rows": [{key: value for key, value in row.items() if key != "payload"} for row in validated],
    }


@router.post("/bulk")
async def import_bulk_employees(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user_id: str | None = Header(None, alias="x-user-id"),
    current_user_email: str | None = Header(None, alias="x-user-email"),
):
    """Import valid rows only, queueing the normal secure activation email for each."""
    actor = require_admin_employee(db, current_user_id, current_user_email)
    content = await file.read()
    rows = _validate_bulk_employee_rows(db, _parse_bulk_employee_file(file.filename or "", content))
    imported = 0
    skipped = []
    for row in rows:
        if not row["valid"]:
            skipped.append({"row": row["row"], "email": row["email"], "error": row["error"]})
            continue
        try:
            data = AddEmployeeRequest(**row["payload"])
            result = create_employee(db, data)
            if not result.success or not result.employee_id:
                skipped.append({"row": row["row"], "email": row["email"], "error": result.message})
                continue
            imported += 1
            log_audit(
                db,
                actor,
                action="employee.created.bulk",
                entity_type="employee",
                entity_id=result.employee_id,
                new_values=data.model_dump(),
                metadata={"source_file": file.filename, "source_row": row["row"]},
            )
            db.commit()
        except Exception as exc:
            db.rollback()
            logger.exception("Bulk employee import failed for row %s", row["row"])
            skipped.append({"row": row["row"], "email": row["email"], "error": str(exc)})
    return {"imported": imported, "skipped": len(skipped), "skipped_rows": skipped}


@router.get("/organization", response_model=dict)
async def organization_chart(
    db: Session = Depends(get_db),
    current_user_id: str | None = Header(None, alias="x-user-id"),
    current_user_email: str | None = Header(None, alias="x-user-email"),
):
    """Return the minimal flat employee graph needed to render reporting lines."""
    actor = get_current_employee(db, current_user_id, current_user_email)
    employees = db.query(Employee).filter(Employee.is_active.is_(True)).order_by(
        Employee.first_name.asc(), Employee.last_name.asc()
    ).all()
    resolved_manager_ids = stabilize_manager_map(
        employees,
        {employee.id: resolve_manager_id(employee, employees) for employee in employees},
    )
    report_counts: dict[str, int] = {}
    for manager_id in resolved_manager_ids.values():
        if manager_id:
            report_counts[manager_id] = report_counts.get(manager_id, 0) + 1

    now = datetime.utcnow()
    rows = []
    for employee in employees:
        last_active = employee.last_active_at
        role_label = (employee.role or "Employee").replace("_", " ").replace("-", " ").title()
        rows.append({
            "id": employee.id,
            "name": employee_name(employee),
            "designation": employee.designation or role_label,
            "department": employee.department or "Not assigned",
            "role": employee.role,
            "reporting_manager_id": resolved_manager_ids[employee.id],
            "reports_count": report_counts.get(employee.id, 0),
            "is_online": bool(last_active and (now - last_active).total_seconds() <= 15 * 60),
            "is_current_user": employee.id == actor.id,
        })

    return {
        "employees": rows,
        "current_user_id": actor.id,
        "employee_count": len(rows),
        "department_count": len({row["department"] for row in rows if row["department"] != "Not assigned"}),
    }


@router.get("/{employee_id}")
async def get_employee(
    employee_id: str,
    db: Session = Depends(get_db),
    current_user_id: str | None = Header(None, alias="x-user-id"),
    current_user_email: str | None = Header(None, alias="x-user-email"),
):
    """Get single employee details."""
    actor = get_current_employee(db, current_user_id, current_user_email)
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    if not can_view_employee_detail(actor, emp):
        raise HTTPException(status_code=403, detail="Not authorized to view this employee.")

    if actor.id != employee_id:
        log_sensitive_access(
            db,
            actor,
            action="employee_profile_view",
            target_type="employee",
            target_id=employee_id,
            sensitivity_level="confidential",
            reason="Admin profile lookup",
        )
        db.commit()

    return serialize_employee(emp)


@router.get("/{employee_id}/preview")
async def get_employee_preview(
    employee_id: str,
    db: Session = Depends(get_db),
    current_user_id: str | None = Header(None, alias="x-user-id"),
    current_user_email: str | None = Header(None, alias="x-user-email"),
):
    """Executive employee preview drawer data."""
    actor = require_admin_employee(db, current_user_id, current_user_email)
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    log_sensitive_access(
        db,
        actor,
        action="employee_preview_view",
        target_type="employee",
        target_id=employee_id,
        sensitivity_level="confidential",
        reason="Admin employee preview",
    )
    db.commit()

    today = date.today()
    current_leave = db.query(LeaveRequest).filter(
        LeaveRequest.employee_id == employee_id,
        LeaveRequest.status == "approved",
        LeaveRequest.start_date <= today,
        LeaveRequest.end_date >= today,
    ).first()
    upcoming_leave = db.query(LeaveRequest).filter(
        LeaveRequest.employee_id == employee_id,
        LeaveRequest.status.in_(["approved", "pending"]),
        LeaveRequest.start_date > today,
    ).order_by(LeaveRequest.start_date.asc()).first()

    leave_balance_query = db.query(
        func.sum(LeaveBalance.total_days - LeaveBalance.used_days)
    ).filter(
        LeaveBalance.employee_id == employee_id,
        LeaveBalance.year == today.year,
    )
    leave_balance = leave_balance_query.scalar()
    has_leave_balance = db.query(LeaveBalance.id).filter(
        LeaveBalance.employee_id == employee_id,
        LeaveBalance.year == today.year,
    ).first() is not None

    active_allocations = db.query(Allocation).filter(
        Allocation.employee_id == employee_id,
        Allocation.status == "active",
        Allocation.start_date <= today,
        or_(Allocation.end_date.is_(None), Allocation.end_date >= today),
    ).all()
    allocation_status = "allocated" if active_allocations else None

    completed_courses = db.query(TrainingEnrollment).filter(
        TrainingEnrollment.employee_id == employee_id,
        TrainingEnrollment.status == "completed",
    ).count()
    total_courses = db.query(TrainingEnrollment).filter(
        TrainingEnrollment.employee_id == employee_id,
    ).count()
    learning_percent = round((completed_courses / total_courses) * 100) if total_courses else 0

    performance = db.query(EmployeePerformanceSnapshot).filter(
        EmployeePerformanceSnapshot.employee_id == employee_id,
    ).order_by(EmployeePerformanceSnapshot.updated_at.desc()).first()

    audit_logs = db.query(EmployeeAuditLog).filter(
        EmployeeAuditLog.employee_id == employee_id,
    ).order_by(EmployeeAuditLog.changed_at.desc()).limit(5).all()

    days_inactive = None
    if emp.last_active_at:
        days_inactive = (datetime.utcnow() - emp.last_active_at).days

    stored_access_level = (emp.access_level or "").strip()
    access_level = stored_access_level if stored_access_level and stored_access_level.lower() != "standard" else emp.role
    mfa_status = "not_available"
    if emp.mfa_enabled or (emp.totp_secret and not emp.is_first_login):
        mfa_status = "enabled"
    elif emp.totp_secret and emp.is_first_login:
        mfa_status = "pending_setup"

    return {
        "employee": serialize_employee(emp),
        "account_activation": {
            "account_status": "pending_activation" if emp.is_first_login else ("active" if emp.is_active else "inactive"),
            "activation_code": emp.setup_code if emp.is_first_login else None,
            "invite_status": "pending" if emp.is_first_login else "accepted",
        },
        "workforce_status": {
            "employment_status": emp.employment_status,
            "availability": "on_leave" if current_leave else None,
            "allocation_status": allocation_status,
            "employment_type": emp.workforce_type,
            "active_allocations": len(active_allocations),
        },
        "last_activity": {
            "last_login_at": str(emp.last_login_at) if emp.last_login_at else None,
            "last_active_at": str(emp.last_active_at) if emp.last_active_at else None,
            "last_active_status": "active_recently" if days_inactive is not None and days_inactive <= 7 else "inactive_or_unknown",
            "days_inactive": days_inactive,
        },
        "leave_summary": {
            "available_leave_days": float(leave_balance) if has_leave_balance and leave_balance is not None else None,
            "current_leave_status": "on_leave" if current_leave else "available",
            "upcoming_leave_start": str(upcoming_leave.start_date) if upcoming_leave else None,
            "upcoming_leave_end": str(upcoming_leave.end_date) if upcoming_leave else None,
            "upcoming_leave_status": upcoming_leave.status if upcoming_leave else None,
        },
        "learning_progress": {
            "completed_courses": completed_courses,
            "total_courses": total_courses,
            "completion_percentage": learning_percent,
        },
        "performance_snapshot": {
            "latest_rating": float(performance.latest_rating) if performance and performance.latest_rating is not None else None,
            "last_review_date": str(performance.last_review_date) if performance and performance.last_review_date else None,
            "kpi_score": float(performance.kpi_score) if performance and performance.kpi_score is not None else None,
        },
        "it_access": {
            "access_level": access_level,
            "mfa_enabled": True if mfa_status == "enabled" else None,
            "mfa_status": mfa_status,
            "assigned_systems_count": len(active_allocations) if active_allocations else None,
            "last_login_at": str(emp.last_login_at) if emp.last_login_at else None,
            "device_tracking_available": False,
            "device_assigned": None,
        },
        "audit_changes": [
            {
                "id": log.id,
                "action_type": log.action_type,
                "field_name": log.field_name,
                "old_value": log.old_value,
                "new_value": log.new_value,
                "changed_by": log.changed_by,
                "changed_at": str(log.changed_at),
            }
            for log in audit_logs
        ],
    }


@router.post("/{employee_id}/remind-emergency-contact")
async def remind_emergency_contact(
    employee_id: str,
    db: Session = Depends(get_db),
    current_user_id: str | None = Header(None, alias="x-user-id"),
    current_user_email: str | None = Header(None, alias="x-user-email"),
):
    """Ask an employee to update missing emergency contact information."""
    actor = require_admin_employee(db, current_user_id, current_user_email)
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")
    if not emp.is_active:
        raise HTTPException(status_code=400, detail="Reminders can only be sent to active employees.")

    emergency_contact_complete = all([
        bool((emp.emergency_contact_name or "").strip()),
        bool((emp.emergency_contact_phone or "").strip()),
        bool((emp.emergency_contact_relation or "").strip()),
    ])
    if emergency_contact_complete:
        db.query(ActionInboxItem).filter(
            ActionInboxItem.assigned_to_user_id == emp.id,
            ActionInboxItem.item_type == "profile_update",
            ActionInboxItem.related_entity_type == "employee",
            ActionInboxItem.related_entity_id == emp.id,
            ActionInboxItem.status == "pending",
        ).update(
            {
                ActionInboxItem.status: "completed",
                ActionInboxItem.updated_at: datetime.utcnow(),
            },
            synchronize_session=False,
        )
        db.query(Notification).filter(
            Notification.user_id == emp.id,
            Notification.notification_type == "profile_update",
            Notification.related_entity_type == "employee",
            Notification.related_entity_id == emp.id,
        ).delete(synchronize_session=False)
        db.commit()
        return {"success": True, "message": "Emergency contact details are already complete."}

    existing_action = db.query(ActionInboxItem).filter(
        ActionInboxItem.assigned_to_user_id == emp.id,
        ActionInboxItem.item_type == "profile_update",
        ActionInboxItem.related_entity_type == "employee",
        ActionInboxItem.related_entity_id == emp.id,
        ActionInboxItem.status == "pending",
    ).first()
    if not existing_action:
        db.add(ActionInboxItem(
            assigned_to_user_id=emp.id,
            item_type="profile_update",
            title="Update emergency contact",
            description="Please add or update your emergency contact details in My Profile.",
            status="pending",
            priority="normal",
            related_entity_type="employee",
            related_entity_id=emp.id,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        ))

    # This is an action the employee must complete, so keep it in Inbox only.
    # Remove older duplicate notification reminders created before this was separated.
    db.query(Notification).filter(
        Notification.user_id == emp.id,
        Notification.notification_type == "profile_update",
        Notification.related_entity_type == "employee",
        Notification.related_entity_id == emp.id,
    ).delete(synchronize_session=False)
    log_sensitive_access(
        db,
        actor,
        action="emergency_contact_reminder_sent",
        target_type="employee",
        target_id=emp.id,
        sensitivity_level="restricted",
        reason="Admin requested employee emergency contact update",
    )
    log_audit(
        db,
        actor,
        action="employee.emergency_contact_reminder_sent",
        entity_type="employee",
        entity_id=emp.id,
        reason="Admin requested employee emergency contact update",
        metadata={"target_employee": f"{emp.first_name} {emp.last_name}".strip()},
    )
    db.commit()
    return {"success": True, "message": "Reminder sent to employee."}


@router.post("/", response_model=AddEmployeeResponse)
async def add_employee(
    data: AddEmployeeRequest,
    db: Session = Depends(get_db),
    current_user_id: str | None = Header(None, alias="x-user-id"),
    current_user_email: str | None = Header(None, alias="x-user-email"),
):
    """Add an employee and queue a secure activation email."""
    actor = require_admin_employee(db, current_user_id, current_user_email)
    try:
        result = create_employee(db, data)
        if result.success and result.employee_id:
            log_audit(
                db,
                actor,
                action="employee.created",
                entity_type="employee",
                entity_id=result.employee_id,
                new_values=data.model_dump(),
                metadata={"security_event": False},
            )
            db.commit()
        return result
    except Exception as e:
        logger.error(f"Error adding employee: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{employee_id}")
async def update_employee(
    employee_id: str,
    data: UpdateEmployeeRequest,
    current_user_id: str | None = Header(None, alias="x-user-id"),
    current_user_email: str | None = Header(None, alias="x-user-email"),
    current_user_name: str | None = Header(None, alias="x-user-name"),
    db: Session = Depends(get_db),
):
    """Update employee details. Current user must be the employee or an admin."""
    actor = get_current_employee(db, current_user_id, current_user_email)
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    is_self = actor.id == employee_id
    is_admin = is_admin_role(actor.role)

    if not (is_self or is_admin):
        log_authorization_failure(
            db,
            actor,
            action="employee.update",
            entity_type="employee",
            entity_id=employee_id,
            reason="User attempted to update an employee they do not own.",
        )
        db.commit()
        raise HTTPException(status_code=403, detail="Not authorized to update this employee")

    updates = data.model_dump(exclude_unset=True)
    change_reason = (updates.pop("change_reason", None) or "").strip()
    if is_self and not is_admin:
        blocked_fields = sorted(set(updates) - SELF_PROFILE_FIELDS)
        if blocked_fields:
            log_authorization_failure(
                db,
                actor,
                action="employee.update_restricted_fields",
                entity_type="employee",
                entity_id=employee_id,
                reason=f"Restricted fields attempted: {', '.join(blocked_fields)}",
            )
            db.commit()
            raise HTTPException(
                status_code=403,
                detail=f"Employees cannot modify restricted profile fields: {', '.join(blocked_fields)}.",
            )
    elif is_admin:
        blocked_fields = sorted(set(updates) - ADMIN_EMPLOYEE_FIELDS)
        if blocked_fields:
            raise HTTPException(
                status_code=403,
                detail=f"These employee fields require a dedicated security workflow: {', '.join(blocked_fields)}.",
            )

    changed_update_fields = {
        field
        for field, value in updates.items()
        if hasattr(emp, field) and getattr(emp, field) != value
    }
    sensitive_changes = changed_update_fields.intersection(SENSITIVE_EMPLOYMENT_FIELDS)
    if is_admin and sensitive_changes and not change_reason:
        raise HTTPException(
            status_code=422,
            detail="A reason is required when changing employment details.",
        )

    if "reporting_manager" in updates:
        requested_manager = (updates.get("reporting_manager") or "").strip()
        if not requested_manager or requested_manager.lower() == "not assigned":
            updates["reporting_manager"] = ""
            updates["manager_id"] = None
        else:
            normalized_manager = requested_manager.lower()
            manager = db.query(Employee).filter(
                or_(
                    func.lower(Employee.work_email) == normalized_manager,
                    func.lower(
                        func.trim(
                            func.coalesce(Employee.first_name, "")
                            + literal(" ")
                            + func.coalesce(Employee.last_name, "")
                        )
                    ) == normalized_manager,
                )
            ).first()
            if not manager:
                raise HTTPException(status_code=422, detail="The selected reporting manager was not found.")
            if manager.id == emp.id:
                raise HTTPException(status_code=422, detail="An employee cannot report to themselves.")
            if normalize_role(manager.role) not in {"manager", "super_admin", "admin", "hr_admin", "global_access"}:
                raise HTTPException(status_code=422, detail="The selected employee is not eligible to be a reporting manager.")
            if creates_reporting_cycle(db, emp.id, manager.id):
                raise HTTPException(status_code=422, detail="This reporting-manager change would create a circular reporting line.")
            updates["reporting_manager"] = employee_name(manager)
            updates["manager_id"] = manager.id

    emergency_contact_fields = (
        "emergency_contact_name",
        "emergency_contact_phone",
        "emergency_contact_relation",
    )
    if any(field in updates for field in emergency_contact_fields):
        emergency_contact_values = {
            "Contact name": (updates.get("emergency_contact_name", emp.emergency_contact_name) or "").strip(),
            "Contact phone": (updates.get("emergency_contact_phone", emp.emergency_contact_phone) or "").strip(),
            "Relationship": (updates.get("emergency_contact_relation", emp.emergency_contact_relation) or "").strip(),
        }
        has_any_emergency_contact = any(emergency_contact_values.values())
        has_all_emergency_contact = all(emergency_contact_values.values())
        if has_any_emergency_contact and not has_all_emergency_contact:
            missing = [label for label, value in emergency_contact_values.items() if not value]
            raise HTTPException(
                status_code=422,
                detail=f"Emergency contact is incomplete. Missing: {', '.join(missing)}.",
            )

    old_values = {field: getattr(emp, field) for field in updates.keys() if hasattr(emp, field)}

    # Apply updates
    for field, value in updates.items():
        if hasattr(emp, field):
            setattr(emp, field, value)

    emp.last_updated_at = datetime.utcnow()
    emp.updated_by = changed_by_value(actor.work_email, current_user_name, actor.id)

    emergency_contact_complete = all([
        bool((emp.emergency_contact_name or "").strip()),
        bool((emp.emergency_contact_phone or "").strip()),
        bool((emp.emergency_contact_relation or "").strip()),
    ])
    if emergency_contact_complete:
        db.query(ActionInboxItem).filter(
            ActionInboxItem.assigned_to_user_id == emp.id,
            ActionInboxItem.item_type == "profile_update",
            ActionInboxItem.related_entity_type == "employee",
            ActionInboxItem.related_entity_id == emp.id,
            ActionInboxItem.status == "pending",
        ).update(
            {
                ActionInboxItem.status: "completed",
                ActionInboxItem.updated_at: datetime.utcnow(),
            },
            synchronize_session=False,
        )
        db.query(Notification).filter(
            Notification.user_id == emp.id,
            Notification.notification_type == "profile_update",
            Notification.related_entity_type == "employee",
            Notification.related_entity_id == emp.id,
        ).delete(synchronize_session=False)

    log_employee_changes(db, employee_id, old_values, updates, emp.updated_by)
    action = "user_profile_updated" if is_self and not is_admin else "employee.updated"
    field_changes = changed_fields(old_values, updates)
    if field_changes:
        log_audit(
            db,
            actor,
            action=action,
            entity_type="employee",
            entity_id=employee_id,
            old_values=old_values,
            new_values=updates,
            changed_fields_payload=field_changes,
            metadata={
                "changed_by": emp.updated_by,
                **({"change_reason": change_reason} if change_reason else {}),
            },
        )

    if is_admin and not is_self and sensitive_changes:
        readable_fields = ", ".join(
            field.replace("_", " ") for field in sorted(sensitive_changes)
        )
        db.add(Notification(
            user_id=emp.id,
            title="Employment details updated",
            message=(
                f"{employee_name(actor)} updated your {readable_fields}. "
                f"Reason: {change_reason}"
            ),
            type="system",
            notification_type="employee_record_updated",
            related_entity_type="employee",
            related_entity_id=emp.id,
            link_url="/profile",
        ))

    db.commit()
    db.refresh(emp)

    return {
        "success": True,
        "message": "Employee updated successfully",
        "id": emp.id,
        "employee": serialize_employee(emp),
    }


@router.post("/{employee_id}/upload-profile-picture")
async def upload_profile_picture(
    employee_id: str,
    file: UploadFile = File(...),
    current_user_id: str | None = Header(None, alias="x-user-id"),
    current_user_email: str | None = Header(None, alias="x-user-email"),
    current_user_name: str | None = Header(None, alias="x-user-name"),
    db: Session = Depends(get_db),
):
    """Upload profile picture for an employee. User can only upload for their own profile or super_admin can upload for anyone."""
    actor = get_current_employee(db, current_user_id, current_user_email)
    emp = db.query(Employee).filter(Employee.id == employee_id).first()
    if not emp:
        raise HTTPException(status_code=404, detail="Employee not found")

    if actor.id != employee_id and not is_admin_role(actor.role):
        raise HTTPException(status_code=403, detail="Not authorized to upload profile picture for this employee")

    # Validate file type
    allowed_types = {"image/jpeg", "image/png", "image/webp", "image/gif"}
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=400,
            detail="Invalid file type. Allowed: JPEG, PNG, WebP, GIF"
        )

    # Validate file size (max 5MB)
    file_content = await file.read()
    if len(file_content) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size exceeds 5MB limit")

    # Convert to base64
    image_base64 = base64.b64encode(file_content).decode('utf-8')
    data_uri = f"data:{file.content_type};base64,{image_base64}"

    # Update employee profile picture
    emp.profile_image_url = data_uri
    emp.last_updated_at = datetime.utcnow()
    emp.updated_by = actor.work_email or current_user_name or actor.id or "unknown"
    log_audit(
        db,
        actor,
        action="employee.profile_picture_uploaded",
        entity_type="employee",
        entity_id=employee_id,
        old_values={"profile_image_url": "[IMAGE_PREVIOUS]"},
        new_values={"profile_image_url": "[IMAGE_UPDATED]"},
        metadata={"content_type": file.content_type, "file_size": len(file_content)},
    )
    db.commit()
    db.refresh(emp)

    return {
        "success": True,
        "message": "Profile picture uploaded successfully",
        "profile_image_url": emp.profile_image_url,
        "employee": serialize_employee(emp),
    }
